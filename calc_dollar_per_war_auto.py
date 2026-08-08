#!/usr/bin/env python3
"""
calc_dollar_per_war_auto.py — compute $/WAR by year from Spotrac + FanGraphs + local xlsx WAR files.

Data sources:
  Spotrac free-agent tracker  — who entered FA each year
  FanGraphs contracts API     — new contract AAV; ContractType filters FA vs Extension/Arb
  Local xlsx files (OneDrive) — fWAR by season (downloaded from fangraphs.com/leaders/major-league)

Year offset note:
  Spotrac year=N lists the FA class that signed AFTER the N season.
  Their FanGraphs contracts have startSeason=N+1 (e.g. Spotrac /2025/ → startSeason=2026).
  WAR is averaged over the N seasons prior to startSeason (the seasons just completed).

Usage:
  python calc_dollar_per_war_auto.py --years 2025
  python calc_dollar_per_war_auto.py --years 2022 2023 2024 2025
  python calc_dollar_per_war_auto.py --years 2025 --war-years 3 --min-aav 8
  python calc_dollar_per_war_auto.py --years 2025 --fg-war-dir "C:/path/to/xlsx/files"
"""

import math
import os
import re
import shutil
import statistics
import tempfile
import time
import argparse
from collections import defaultdict
from pathlib import Path

import requests
from bs4 import BeautifulSoup
import openpyxl
import pybaseball

pybaseball.cache.enable()

FG_WAR_DIR_DEFAULT = str(Path.home() / "fwar_data")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept": "application/json, */*",
    "Referer": "https://www.fangraphs.com/",
}
FG_CONTRACTS = "https://www.fangraphs.com/api/roster-resource/contracts/player?playerid={}"
RATE_LIMIT_S = 0.5


# ---------------------------------------------------------------------------
# WAR cache — local xlsx files (fWAR)
# ---------------------------------------------------------------------------

def build_war_cache(years: list, n_prior: int, fg_war_dir: str) -> dict:
    """
    Load fWAR from local xlsx files exported from FanGraphs.
    Returns {(mlbam_id, season): fWAR}.

    Expected filename patterns (either batting or hitting, both checked):
      {season}_batting_war_csv.xlsx
      {season}_hitting_war_csv.xlsx
      {season}_pitching_war_csv.xlsx
    """
    # Need seasons from (min_year - n_prior) to max_year inclusive.
    # +1 because startSeason can be year+1, so WAR avg starts at year (not year-1).
    min_season = min(years) - n_prior
    max_season = max(years)
    needed = set(range(min_season, max_season + 1))

    cache = {}
    loaded = []

    for filename in sorted(os.listdir(fg_war_dir)):
        m = re.match(r"^(\d{4})_+(batting|hitting|pitching)_war", filename)
        if not m or not filename.endswith(".xlsx"):
            continue
        season = int(m.group(1))
        if season not in needed:
            continue

        path = os.path.join(fg_war_dir, filename)
        # Copy to temp to avoid OneDrive sync locks
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        shutil.copy2(path, tmp.name)
        wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(max_row=1))]
        war_idx   = headers.index("WAR")
        mlbam_idx = headers.index("MLBAMID")
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            mlbam = row[mlbam_idx]
            war   = row[war_idx]
            if mlbam is None or war is None:
                continue
            war_val = float(war)
            if season == 2020:
                war_val *= 162 / 60  # prorate 60-game season to full-season equivalent
            key = (int(mlbam), season)
            cache[key] = max(cache.get(key, float("-inf")), war_val)
            count += 1
        wb.close()
        os.unlink(tmp.name)
        loaded.append(f"{filename} ({count} rows)")

    for f in loaded:
        print(f"  Loaded: {f}")
    return cache


# ---------------------------------------------------------------------------
# Spotrac + FanGraphs fetchers
# ---------------------------------------------------------------------------

def get_fa_class(year: int) -> list:
    """Scrape Spotrac FA tracker. Returns [(player_name, fa_type), ...]."""
    r = requests.get(
        f"https://www.spotrac.com/mlb/free-agents/{year}/",
        headers=HEADERS, timeout=15,
    )
    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("table")
    players = []
    for tr in table.find("tbody").find_all("tr"):
        cells = [td.get_text(strip=True) for td in tr.find_all("td")]
        if len(cells) >= 6:
            name = re.sub(r"QOI.*|[0-9]+-DAY.*", "", cells[0]).strip()
            fa_type = cells[5]
            players.append((name, fa_type))
    return players


def get_player_ids(player_name: str, active_since: int):
    """Return (fg_id, mlbam_id). Either can be None on failure."""
    parts = player_name.strip().split()
    if len(parts) < 2:
        return None, None
    first = parts[0]
    last  = " ".join(parts[1:])

    for fuzzy in (False, True):
        try:
            lkp = pybaseball.playerid_lookup(last, first, fuzzy=fuzzy)
            lkp = lkp[lkp["key_fangraphs"].notna()]
            lkp = lkp[lkp["mlb_played_last"] >= active_since - 2]
            if len(lkp):
                row     = lkp.iloc[0]
                fg_val  = row["key_fangraphs"]
                mlb_val = row["key_mlbam"]
                fg_id   = int(fg_val)  if not math.isnan(float(fg_val))  else None
                mlb_id  = int(mlb_val) if not math.isnan(float(mlb_val)) else None
                return fg_id, mlb_id
        except Exception:
            pass
    return None, None


def get_fg_contract_aav(fg_id: int, spotrac_year: int):
    """
    Return (aav_m, start_season) for the FA contract signed after spotrac_year, or None.

    Tries startSeason == spotrac_year+1 first (the normal case: Spotrac /N/ → FG startSeason N+1),
    then falls back to startSeason == spotrac_year for same-year signings.
    """
    time.sleep(RATE_LIMIT_S)
    r = requests.get(FG_CONTRACTS.format(fg_id), headers=HEADERS, timeout=10)
    if r.status_code != 200:
        return None
    contracts = r.json()
    for target_start in (spotrac_year + 1, spotrac_year):
        for contract in contracts:
            cs = contract.get("contractSummary", {})
            if (
                cs.get("ContractType") == "Free Agent"
                and cs.get("startSeason") == target_start
                and cs.get("AAV")
            ):
                return cs["AAV"] / 1_000_000, target_start
    return None


def avg_war(mlbam_id: int, start_season: int, cache: dict, n_prior: int):
    """
    Average fWAR over n_prior seasons before start_season (i.e. start_season-1 back).
    Skips seasons with WAR <= 0 (injury, minors).
    """
    vals = [
        cache[(mlbam_id, start_season - i)]
        for i in range(1, n_prior + 1)
        if cache.get((mlbam_id, start_season - i), 0) > 0
    ]
    return (sum(vals) / len(vals), len(vals)) if vals else None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Compute $/WAR by FA class year")
    ap.add_argument("--years",       type=int, nargs="+", default=[2025])
    ap.add_argument("--war-years",   type=int, default=4, dest="war_years",
                    help="Prior seasons to average fWAR over (default 4)")
    ap.add_argument("--min-aav",     type=float, default=5.0, dest="min_aav",
                    help="Min new contract AAV $M to include (default 5)")
    ap.add_argument("--min-war",     type=float, default=1.0, dest="min_war",
                    help="Min avg WAR to include — filters near-zero WAR outliers (default 1.0)")
    ap.add_argument("--fg-war-dir",  default=FG_WAR_DIR_DEFAULT, dest="fg_war_dir",
                    help="Directory containing {year}_{batting|pitching}_war_csv.xlsx files")
    args = ap.parse_args()

    print(f"\nLoading fWAR from xlsx files in: {args.fg_war_dir}")
    war_cache = build_war_cache(args.years, args.war_years, args.fg_war_dir)
    print(f"  {len(war_cache)} player-seasons cached\n")

    by_year = defaultdict(list)
    skipped = []

    for year in sorted(args.years):
        print(f"=== {year} FA class (Spotrac /mlb/free-agents/{year}/) ===")
        fa_class = get_fa_class(year)
        print(f"  {len(fa_class)} players in FA class\n")

        for player_name, fa_type in fa_class:
            fg_id, mlbam_id = get_player_ids(player_name, year)
            if fg_id is None:
                skipped.append((year, player_name, "no FG ID"))
                continue
            if mlbam_id is None:
                skipped.append((year, player_name, "no MLBAM ID"))
                continue

            result = get_fg_contract_aav(fg_id, year)
            if result is None:
                skipped.append((year, player_name, f"no FA contract near {year}"))
                continue
            aav, start_season = result
            if aav < args.min_aav:
                continue

            war_result = avg_war(mlbam_id, start_season, war_cache, args.war_years)
            if war_result is None:
                skipped.append((year, player_name, "no fWAR data"))
                continue

            avg_w, n_yrs = war_result
            if avg_w < args.min_war:
                skipped.append((year, player_name, f"avg fWAR {avg_w:.2f} below --min-war {args.min_war}"))
                continue
            dpw = aav / avg_w
            by_year[year].append((player_name, aav, avg_w, n_yrs, dpw))
            print(f"  {player_name:<28} AAV=${aav:>5.1f}M  fWAR={avg_w:.2f} ({n_yrs}yr)"
                  f"  $/fWAR=${dpw:.2f}M  [startSeason={start_season}]")

        print()

    print(f"\n{'Year':>4}   {'N':>3}   {'Median $/fWAR':>13}   {'Mean $/fWAR':>11}")
    print("-" * 48)
    results = {}
    for year in sorted(by_year):
        vals = [x[4] for x in by_year[year]]
        med  = statistics.median(vals)
        mean = statistics.mean(vals)
        results[year] = round(med, 1)
        print(f"  {year:>4}   {len(vals):>3}   ${med:>11.2f}M   ${mean:>9.2f}M")

    print("\n# Paste into DOLLAR_PER_WAR_BY_YEAR in comps.py:")
    print("DOLLAR_PER_WAR_BY_YEAR = {" + ", ".join(f"{y}: {v}" for y, v in sorted(results.items())) + "}")

    if skipped:
        print(f"\n  {len(skipped)} players skipped (first 40):")
        for yr, name, reason in skipped[:40]:
            print(f"    {yr}  {name:<28}  {reason}")


if __name__ == "__main__":
    main()
